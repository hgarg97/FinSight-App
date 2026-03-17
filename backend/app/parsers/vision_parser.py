"""Vision LLM parser — uses OpenAI gpt-4o to extract transactions from Excel and PDF files.

Excel: read raw cell data with openpyxl → pass as formatted text to gpt-4o.
PDF:   render each page to a PNG image with PyMuPDF → pass as base64 images to gpt-4o vision.
"""

import base64
import json
import logging
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from .base_parser import BaseParser
from .normalizer import normalize_merchant
from ..schemas.transaction import TransactionCreate

logger = logging.getLogger(__name__)

_SYSTEM_PROMPT = """You are a financial data extraction assistant. Your job is to parse bank statements, credit card statements, and financial documents and extract every transaction.

Return ONLY a valid JSON array — no prose, no markdown code fences, no explanation.

Each element must be an object with these fields:
- "date": string in YYYY-MM-DD format
- "merchant": string (raw merchant/description name from the document)
- "amount": number (positive = money received/credit, negative = money spent/debit)
- "category": string or null (infer if obvious, otherwise null)

Rules:
- Skip summary rows (opening balance, closing balance, total, subtotal).
- Skip duplicate header rows.
- If only a debit column exists, amount is negative; if only a credit column, amount is positive.
- If separate debit and credit columns exist, use the non-zero one with correct sign.
- Do not invent transactions that are not in the document.
- Return [] if no transactions are found.
"""


def _call_openai_text(content: str) -> str:
    """Send text content to OpenAI and return raw response text."""
    from openai import OpenAI
    from ..config import settings

    client = OpenAI(api_key=settings.OPENAI_API_KEY)
    response = client.chat.completions.create(
        model=settings.LLM_MODEL,
        max_tokens=8192,
        messages=[
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user", "content": content},
        ],
    )
    return response.choices[0].message.content or ""


def _call_openai_vision(image_blocks: List[Dict[str, Any]], extra_text: str = "") -> str:
    """Send one or more base64 PNG images to OpenAI vision and return raw response text."""
    from openai import OpenAI
    from ..config import settings

    client = OpenAI(api_key=settings.OPENAI_API_KEY)

    user_content: List[Dict] = []
    for block in image_blocks:
        user_content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/png;base64,{block['data']}",
                "detail": "high",
            },
        })
    user_content.append({
        "type": "text",
        "text": extra_text or "Extract all financial transactions from this bank/credit card statement.",
    })

    response = client.chat.completions.create(
        model=settings.LLM_MODEL,
        max_tokens=8192,
        messages=[
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user", "content": user_content},
        ],
    )
    return response.choices[0].message.content or ""


def _parse_llm_response(raw: str) -> List[Dict[str, Any]]:
    """Extract and parse the JSON array from the LLM response."""
    text = raw.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    text = text.strip()

    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1:
        logger.warning("LLM response contained no JSON array")
        return []

    try:
        return json.loads(text[start : end + 1])
    except json.JSONDecodeError as exc:
        logger.error("Failed to parse LLM JSON response: %s\n%s", exc, text[:500])
        return []


def _parse_date_str(s: Any) -> Optional[date]:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _records_to_transactions(records: List[Dict[str, Any]]) -> List[TransactionCreate]:
    transactions: List[TransactionCreate] = []
    for rec in records:
        try:
            txn_date = _parse_date_str(rec.get("date"))
            if txn_date is None:
                continue

            raw_merchant = str(rec.get("merchant", "")).strip()
            if not raw_merchant:
                continue

            amount = float(rec.get("amount", 0))
            if amount == 0:
                continue

            merchant = normalize_merchant(raw_merchant)
            category = rec.get("category") or None
            txn_type = "income" if amount > 0 else "expense"

            transactions.append(
                TransactionCreate(
                    date=txn_date,
                    merchant=merchant,
                    amount_total=abs(amount),
                    transaction_type=txn_type,
                    category=str(category) if category else None,
                )
            )
        except Exception as exc:
            logger.warning("Skipping record due to error: %s | record: %s", exc, rec)

    return transactions


# ── Excel parser ───────────────────────────────────────────────────────────────

class ExcelVisionParser(BaseParser):
    """Extracts transactions from .xlsx/.xls files using OpenAI (text mode)."""

    def parse(self, file_path: str, user_id: str) -> List[TransactionCreate]:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel parsing. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Pick sheet with most rows
        best_sheet = max(wb.worksheets, key=lambda ws: ws.max_row)
        ws = best_sheet

        rows_text: List[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if not any(c.strip() for c in cells):
                continue
            rows_text.append("\t".join(cells))

        if not rows_text:
            logger.warning("Excel file %s appears to be empty", file_path)
            return []

        table_text = "\n".join(rows_text)
        prompt = (
            f"Below is the raw cell data extracted from an Excel bank/credit card statement "
            f"(sheet: '{best_sheet.title}'). Tabs separate columns.\n\n"
            f"{table_text}\n\n"
            "Extract all financial transactions and return as JSON."
        )

        logger.info("Sending Excel data to OpenAI (%d rows)…", len(rows_text))
        raw = _call_openai_text(prompt)
        records = _parse_llm_response(raw)
        logger.info("OpenAI extracted %d records from Excel", len(records))

        return _records_to_transactions(records)


# ── PDF parser ─────────────────────────────────────────────────────────────────

class PDFVisionParser(BaseParser):
    """Extracts transactions from PDF files using OpenAI vision (page images)."""

    MAX_PAGES = 10  # guard against huge PDFs

    def parse(self, file_path: str, user_id: str) -> List[TransactionCreate]:
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise RuntimeError("PyMuPDF is required for PDF parsing. Run: pip install PyMuPDF")

        doc = fitz.open(file_path)
        num_pages = min(len(doc), self.MAX_PAGES)

        if num_pages == 0:
            logger.warning("PDF %s has no pages", file_path)
            return []

        image_blocks: List[Dict[str, Any]] = []
        for page_num in range(num_pages):
            page = doc[page_num]
            # 150 DPI — good balance of quality vs payload size
            mat = fitz.Matrix(150 / 72, 150 / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            png_bytes = pix.tobytes("png")
            image_blocks.append({"data": base64.b64encode(png_bytes).decode()})

        doc.close()

        extra_text = (
            f"This is a bank/credit card statement PDF ({num_pages} page(s) shown). "
            "Extract all financial transactions and return as JSON."
        )

        logger.info("Sending %d PDF page(s) to OpenAI vision…", num_pages)
        raw = _call_openai_vision(image_blocks, extra_text)
        records = _parse_llm_response(raw)
        logger.info("OpenAI extracted %d records from PDF", len(records))

        return _records_to_transactions(records)
